import csv
import openpyxl
import os

def csv_to_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active

    with open('C:\\Users\\a\\Desktop\\PTSD_03162018.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save('C:\\Users\\a\\Desktop\\file.xlsx')

def remove_NaN():
    orig_name = 'C:\\Users\\a\\Desktop\\test.xlsx'
    edited_name = 'C:\\Users\\a\\Desktop\\test_edited.xlsx'
    orig_wb = openpyxl.load_workbook(orig_name)
    orig_wb.save(edited_name)
    wb = openpyxl.load_workbook(edited_name)

    ws = wb.active
    max_cells = ws.max_row

    interval_type_active = "ACTIVE"
    interval_type_daily = "DAILY"
    interval_column = 5
    nan_column = 23

    row_removal_list = []

    for i in range(1, max_cells):
        if ws.cell(row=i, column=interval_column).value == interval_type_active:
            if ws.cell(row=i, column=nan_column).value == "NaN":
                ws.cell(row=i, column=nan_column).value = "test"
                row_removal_list.append(i)
                ws.delete_rows(i, amount=1)
                print(i)

    wb.save(edited_name)